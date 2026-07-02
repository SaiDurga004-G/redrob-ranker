#!/usr/bin/env python3
"""
Optional: join submission.csv back against candidates.jsonl to produce a
human-readable .xlsx shortlist for internal review (recruiters, teammates).

This is NOT part of the required hackathon submission (that's the CSV
produced by rank.py) -- it's a convenience export.

Usage:
    python scripts/export_xlsx.py \
        --candidates ./candidates.jsonl \
        --submission ./submission.csv \
        --out ./shortlist.xlsx
"""

import argparse
import csv
import gzip
import json
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))


def _open(path):
    if path.endswith(".gz"):
        return gzip.open(path, "rt", encoding="utf-8")
    return open(path, "r", encoding="utf-8")


def load_submission(path):
    with open(path, "r", encoding="utf-8") as f:
        return {row["candidate_id"]: row for row in csv.DictReader(f)}


def load_candidate_lookup(path, wanted_ids):
    lookup = {}
    with _open(path) as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            cid_pos = line.find('"candidate_id"')
            if cid_pos == -1:
                continue
            cand = json.loads(line)
            if cand["candidate_id"] in wanted_ids:
                lookup[cand["candidate_id"]] = cand
            if len(lookup) == len(wanted_ids):
                break
    return lookup


HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Arial", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")

COLUMNS = [
    ("Rank", 6),
    ("Candidate ID", 14),
    ("Score", 8),
    ("Name (anonymized)", 18),
    ("Current Title", 26),
    ("Current Company", 18),
    ("Industry", 16),
    ("Years Exp", 10),
    ("Location", 22),
    ("Notice (days)", 12),
    ("Open to Work", 12),
    ("Recruiter Response Rate", 12),
    ("Last Active", 12),
    ("Top Skills", 30),
    ("Reasoning", 55),
]


def build_workbook(sub_rows, lookup, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Top 100 Shortlist"

    for col_idx, (name, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"

    ordered = sorted(sub_rows.values(), key=lambda r: int(r["rank"]))
    for row_idx, sub in enumerate(ordered, start=2):
        cand = lookup.get(sub["candidate_id"], {})
        profile = cand.get("profile", {})
        signals = cand.get("redrob_signals", {})
        skills = cand.get("skills", [])
        top_skills = ", ".join(s["name"] for s in skills[:6])

        values = [
            int(sub["rank"]),
            sub["candidate_id"],
            float(sub["score"]),
            profile.get("anonymized_name", ""),
            profile.get("current_title", ""),
            profile.get("current_company", ""),
            profile.get("current_industry", ""),
            profile.get("years_of_experience", ""),
            profile.get("location", ""),
            signals.get("notice_period_days", ""),
            "Yes" if signals.get("open_to_work_flag") else "No",
            signals.get("recruiter_response_rate", ""),
            signals.get("last_active_date", ""),
            top_skills,
            sub.get("reasoning", ""),
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = BODY_FONT
            if col_idx == len(COLUMNS):  # reasoning column
                cell.alignment = WRAP

    wb.save(out_path)


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--candidates", required=True)
    parser.add_argument("--submission", required=True)
    parser.add_argument("--out", required=True)
    args = parser.parse_args()

    sub_rows = load_submission(args.submission)
    lookup = load_candidate_lookup(args.candidates, set(sub_rows.keys()))
    build_workbook(sub_rows, lookup, args.out)
    print(f"Wrote {len(sub_rows)} candidates to {args.out}")


if __name__ == "__main__":
    main()
